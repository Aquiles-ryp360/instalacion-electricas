#!/usr/bin/env python3
"""Scraper de catalogo.win para enriquecer BOMs con codigo y nombre SIGA/MEF."""

from __future__ import annotations

import argparse
import csv
import html
import json
import re
import sys
import time
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from rich.console import Console
from rich.progress import Progress, SpinnerColumn, TextColumn, TimeElapsedColumn
from rich.table import Table


COTIZACION_DIR = Path(__file__).resolve().parents[1]
if str(COTIZACION_DIR) not in sys.path:
    sys.path.insert(0, str(COTIZACION_DIR))

try:
    from normalizador_materiales import normalizar_nombre_material
except Exception:  # pragma: no cover - el scraper debe poder correr aislado.
    normalizar_nombre_material = None


BASE_URL = "https://catalogo.win"
SEARCH_ENDPOINTS = ("/api/search", "/search")
DEFAULT_KEY_CANDIDATES = ("item", "nombre", "descripcion")
REQUEST_TIMEOUT = 20
DELAY_SECONDS = 1.5
OUTPUT_FIELDS = [
    "index",
    "input_key",
    "query",
    "query_usada",
    "query_intentos",
    "codigo",
    "nombre",
    "catalogo_id",
    "codigo_osce",
    "codigo_onu",
    "codigo_searchable",
    "unidad_medida",
    "tipo_bien",
    "precio_ref",
    "nombre_grupo",
    "nombre_clase",
    "nombre_familia",
    "fecha_alta",
    "total_hits",
    "processing_time_ms",
    "status",
    "error",
]

console = Console()


class CriticalScraperError(RuntimeError):
    """Error que impide continuar la ejecucion completa del scraper."""


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Busca materiales de un JSON en catalogo.win y guarda el primer "
            "codigo/nombre encontrado por item."
        )
    )
    parser.add_argument("--input", required=True, help="Ruta al JSON de entrada.")
    parser.add_argument("--output", required=True, help="Ruta de salida .csv o .json.")
    parser.add_argument(
        "--key",
        help=(
            "Llave del JSON que contiene el nombre del material. Si se omite, "
            "se intenta item, nombre o descripcion."
        ),
    )
    return parser.parse_args()


def load_json(path: Path) -> Any:
    try:
        with path.open("r", encoding="utf-8") as fh:
            return json.load(fh)
    except FileNotFoundError as exc:
        raise CriticalScraperError(f"No existe el archivo de entrada: {path}") from exc
    except json.JSONDecodeError as exc:
        raise CriticalScraperError(f"JSON invalido en {path}: {exc}") from exc
    except OSError as exc:
        raise CriticalScraperError(f"No se pudo leer {path}: {exc}") from exc


def unwrap_items(raw: Any) -> List[Dict[str, Any]]:
    """Acepta una lista directa o contenedores comunes como materiales/items/data."""
    if isinstance(raw, list):
        items = raw
    elif isinstance(raw, dict):
        for key in ("materiales", "items", "data", "rows"):
            value = raw.get(key)
            if isinstance(value, list):
                items = value
                break
        else:
            items = [raw]
    else:
        raise CriticalScraperError("El JSON debe ser una lista o un objeto con materiales/items/data.")

    normalized: List[Dict[str, Any]] = []
    for index, item in enumerate(items, start=1):
        if isinstance(item, dict):
            normalized.append(item)
        else:
            normalized.append({"item": str(item), "_source_index": index})
    return normalized


def find_material_name(item: Dict[str, Any], requested_key: Optional[str]) -> Tuple[Optional[str], str]:
    keys: Sequence[str] = (requested_key,) if requested_key else DEFAULT_KEY_CANDIDATES
    for key in keys:
        if not key:
            continue
        value = item.get(key)
        if value is not None and str(value).strip():
            return str(value).strip(), key
    return None, requested_key or "|".join(DEFAULT_KEY_CANDIDATES)


def strip_html(value: Any) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"<[^>]+>", "", text)
    return html.unescape(text).strip()


def extract_csrf_token(page_html: str) -> Optional[str]:
    patterns = (
        r'<meta[^>]+name=["\']csrf-token["\'][^>]+content=["\']([^"\']+)["\']',
        r'<meta[^>]+content=["\']([^"\']+)["\'][^>]+name=["\']csrf-token["\']',
        r"CSRF_TOKEN\s*=\s*['\"]([^'\"]+)['\"]",
        r"csrf_token['\"]?\s*[:=]\s*['\"]([^'\"]+)['\"]",
    )
    for pattern in patterns:
        match = re.search(pattern, page_html, flags=re.IGNORECASE)
        if match:
            return match.group(1)
    return None


def decode_flask_cookie_token(cookie_value: str) -> Optional[str]:
    """Fallback best-effort: lee el primer segmento base64 de una cookie Flask firmada."""
    import base64

    if not cookie_value:
        return None
    first_segment = cookie_value.split(".", 1)[0]
    padding = "=" * (-len(first_segment) % 4)
    try:
        decoded = base64.urlsafe_b64decode(first_segment + padding).decode("utf-8")
        data = json.loads(decoded)
    except Exception:
        return None
    token = data.get("csrf_token")
    return str(token) if token else None


def bootstrap_session() -> Tuple[requests.Session, str]:
    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": (
                "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
                "(KHTML, like Gecko) Chrome/120.0 Safari/537.36"
            ),
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "es-PE,es;q=0.9,en;q=0.8",
        }
    )

    try:
        response = session.get(BASE_URL + "/", timeout=REQUEST_TIMEOUT)
        response.raise_for_status()
    except requests.RequestException as exc:
        raise CriticalScraperError(f"No se pudo iniciar sesion contra {BASE_URL}: {exc}") from exc

    token = extract_csrf_token(response.text)
    if not token:
        token = decode_flask_cookie_token(session.cookies.get("session", ""))
    if not token:
        raise CriticalScraperError("No se encontro token CSRF en HTML ni cookie de sesion.")
    if not session.cookies:
        raise CriticalScraperError("La respuesta inicial no entrego cookies de sesion.")
    return session, token


def first_hit_from_payload(payload: Any) -> Optional[Dict[str, Any]]:
    if isinstance(payload, dict):
        for key in ("hits", "results", "items", "data"):
            value = payload.get(key)
            if isinstance(value, list) and value:
                return value[0] if isinstance(value[0], dict) else {"nombre_item": str(value[0])}
        if any(key in payload for key in ("codigo_display", "codigo", "nombre_item", "nombre", "descripcion")):
            return payload
        return None
    if isinstance(payload, list) and payload:
        return payload[0] if isinstance(payload[0], dict) else {"nombre_item": str(payload[0])}
    return None


def query_catalogo_basica(query: str) -> str:
    text = query.strip()
    text = re.sub(r"\s*-\s*C\d+\b", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\bC\d+\b", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\bITM\b", "interruptor termomagnetico", text, flags=re.IGNORECASE)
    text = re.sub(r"\bTCs?\b", "tomacorriente", text, flags=re.IGNORECASE)
    text = text.replace("TW THW", "THW").replace("THW TW", "THW")
    text = re.sub(r"\((?:alimentador|general|trifasico|fase|neutro|tierra)[^)]+\)", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+", " ", text).strip(" -")
    return text


def query_catalogo_especifica(query: str) -> Optional[str]:
    text = query_catalogo_basica(query).lower()
    seccion = re.search(r"(\d+(?:[.,]\d+)?)\s*mm\s*2|\b(\d+(?:[.,]\d+)?)\s*mm2\b", text)
    diametro = re.search(r"(\d+(?:[.,]\d+)?)\s*mm\b", text)
    amperaje = re.search(r"(\d+(?:[.,]\d+)?)\s*a\b", text)

    if "cable" in text or "conductor" in text:
        if seccion:
            valor = next(group for group in seccion.groups() if group)
            return f"cable electrico thw {valor.replace(',', '.')} mm2 cobre"
        return "cable electrico cobre"
    if "tubo" in text and "pvc" in text:
        if diametro:
            return f"tubo pvc sap {diametro.group(1).replace(',', '.')} mm"
        return "tubo pvc sap"
    if "interruptor termomagnetico" in text:
        if amperaje:
            return f"interruptor termomagnetico {amperaje.group(1).replace(',', '.')}A"
        return "interruptor termomagnetico"
    if "diferencial" in text:
        if amperaje:
            return f"interruptor diferencial {amperaje.group(1).replace(',', '.')}A 30mA"
        return "interruptor diferencial 30mA"
    if "tablero" in text:
        polos = re.search(r"(\d+)\s*(?:circuitos|polos)", text)
        if polos:
            return f"tablero electrico {polos.group(1)} polos"
        return "tablero electrico"
    return None


def query_catalogo_genericas(query: str) -> List[str]:
    text = query_catalogo_basica(query).lower()
    queries: List[str] = []

    if "diferencial" in text:
        queries.append("interruptor diferencial")
    if "caja" in text:
        if "octogonal" in text:
            queries.append("caja octogonal")
        if "rectangular" in text:
            if "pase" in text:
                queries.append("caja de pase rectangular")
            queries.append("caja rectangular")
        if "estanca" in text or "hermetica" in text or "hermética" in text:
            queries.append("caja hermetica")
    if "luminaria" in text:
        if "plafon" in text or "plafón" in text:
            queries.append("luminaria plafon led")
        queries.append("luminaria led")
    if "varilla" in text and "tierra" in text:
        queries.append("varilla cobre")
        queries.append("varilla puesta a tierra")
    if "electrodo" in text and "tierra" in text:
        queries.append("electrodo puesta a tierra")
    if "cinta" in text and ("aislar" in text or "aislante" in text):
        queries.append("cinta aislante")
    if "tomacorriente" in text:
        if "tierra" in text:
            queries.append("tomacorriente doble tierra")
            queries.append("tomacorriente con puesta a tierra")
        if "doble" in text:
            queries.append("tomacorriente bipolar doble")
        queries.append("tomacorriente")
    if "interruptor simple" in text:
        queries.append("interruptor simple")
    if "conmut" in text:
        queries.append("interruptor conmutador")

    return queries


def build_query_candidates(query: str) -> List[str]:
    candidates: List[str] = []
    values: List[Optional[str]] = [
        query,
        query_catalogo_basica(query),
        normalizar_nombre_material(query) if normalizar_nombre_material else None,
        query_catalogo_especifica(query),
    ]
    values.extend(query_catalogo_genericas(query))
    for value in values:
        if not value:
            continue
        cleaned = re.sub(r"\s+", " ", str(value)).strip()
        if cleaned and cleaned.lower() not in {c.lower() for c in candidates}:
            candidates.append(cleaned)
    return candidates


def extract_hit_fields(hit: Dict[str, Any]) -> Dict[str, Any]:
    code = (
        hit.get("codigo_display")
        or hit.get("codigo")
        or hit.get("codigo_osce")
        or hit.get("codigo_searchable")
        or hit.get("id")
        or ""
    )
    name = (
        hit.get("nombre_item")
        or hit.get("nombre")
        or hit.get("descripcion")
        or hit.get("name")
        or ""
    )
    formatted = hit.get("_formatted")
    if isinstance(formatted, dict):
        code = code or formatted.get("codigo_display") or formatted.get("codigo") or ""
        name = name or formatted.get("nombre_item") or formatted.get("nombre") or ""
    return {
        "codigo": strip_html(code),
        "nombre": strip_html(name),
        "catalogo_id": strip_html(hit.get("id")),
        "codigo_osce": strip_html(hit.get("codigo_osce")),
        "codigo_onu": strip_html(hit.get("codigo_onu")),
        "codigo_searchable": strip_html(hit.get("codigo_searchable")),
        "unidad_medida": strip_html(hit.get("unidad_medida")),
        "tipo_bien": strip_html(hit.get("tipo_bien")),
        "precio_ref": hit.get("precio_ref", ""),
        "nombre_grupo": strip_html(hit.get("nombre_grupo")),
        "nombre_clase": strip_html(hit.get("nombre_clase")),
        "nombre_familia": strip_html(hit.get("nombre_fam") or hit.get("nombre_familia")),
        "fecha_alta": strip_html(hit.get("fecha_alta")),
    }


def search_catalog(session: requests.Session, csrf_token: str, query: str) -> Dict[str, Any]:
    headers = {
        "Accept": "application/json",
        "X-CSRF-Token": csrf_token,
        "Referer": BASE_URL + "/",
    }
    params = {"q": query, "page": 1, "per_page": 10}
    last_error = ""

    for endpoint in SEARCH_ENDPOINTS:
        url = BASE_URL + endpoint
        try:
            response = session.get(url, params=params, headers=headers, timeout=REQUEST_TIMEOUT)
        except requests.RequestException as exc:
            last_error = str(exc)
            continue

        if response.status_code == 404:
            last_error = "HTTP 404"
            continue
        if response.status_code >= 400:
            raise RuntimeError(f"HTTP {response.status_code}: {response.text[:160]}")
        try:
            return response.json()
        except ValueError as exc:
            last_error = f"Respuesta no JSON desde {url}: {exc}"
            continue

    raise RuntimeError(last_error or "No se obtuvo respuesta JSON de ningun endpoint.")


def build_result_row(index: int, item: Dict[str, Any], key_used: str, query: str) -> Dict[str, Any]:
    return {
        "index": index,
        "input_key": key_used,
        "query": query,
        "query_usada": "",
        "query_intentos": "",
        "codigo": "",
        "nombre": "",
        "catalogo_id": "",
        "codigo_osce": "",
        "codigo_onu": "",
        "codigo_searchable": "",
        "unidad_medida": "",
        "tipo_bien": "",
        "precio_ref": "",
        "nombre_grupo": "",
        "nombre_clase": "",
        "nombre_familia": "",
        "fecha_alta": "",
        "total_hits": "",
        "processing_time_ms": "",
        "status": "pendiente",
        "error": "",
    }


def scrape_items(items: Iterable[Dict[str, Any]], requested_key: Optional[str]) -> List[Dict[str, Any]]:
    session, csrf_token = bootstrap_session()
    rows: List[Dict[str, Any]] = []
    items_list = list(items)

    console.log(
        f"[green]Sesion iniciada[/green]: {len(session.cookies)} cookie(s), token CSRF capturado."
    )

    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        TimeElapsedColumn(),
        console=console,
    ) as progress:
        task = progress.add_task("Consultando catalogo.win", total=len(items_list))
        for index, item in enumerate(items_list, start=1):
            query, key_used = find_material_name(item, requested_key)
            if not query:
                rows.append(
                    {
                "index": index,
                "input_key": key_used,
                "query": "",
                "query_usada": "",
                "query_intentos": "",
                "codigo": "",
                "nombre": "",
                "catalogo_id": "",
                "codigo_osce": "",
                "codigo_onu": "",
                "codigo_searchable": "",
                "unidad_medida": "",
                "tipo_bien": "",
                "precio_ref": "",
                "nombre_grupo": "",
                "nombre_clase": "",
                "nombre_familia": "",
                "fecha_alta": "",
                "total_hits": "",
                "processing_time_ms": "",
                "status": "sin_nombre",
                "error": "No se encontro una llave de material usable.",
            }
                )
                console.log(f"[yellow]Advertencia[/yellow] item {index}: sin nombre de material.")
                progress.advance(task)
                time.sleep(DELAY_SECONDS)
                continue

            row = build_result_row(index, item, key_used, query)
            try:
                candidates = build_query_candidates(query)
                row["query_intentos"] = " | ".join(candidates)
                for candidate in candidates:
                    payload = search_catalog(session, csrf_token, candidate)
                    if isinstance(payload, dict):
                        row["total_hits"] = payload.get("totalHits", "")
                        row["processing_time_ms"] = payload.get("processingTimeMs", "")
                    hit = first_hit_from_payload(payload)
                    if hit:
                        row["query_usada"] = candidate
                        row.update(extract_hit_fields(hit))
                        row["status"] = "ok" if row["codigo"] or row["nombre"] else "sin_codigo_nombre"
                        break
                else:
                    row["status"] = "sin_resultados"
            except Exception as exc:
                row["status"] = "error_item"
                row["error"] = str(exc)
                console.log(f"[yellow]Advertencia[/yellow] item {index} ({query}): {exc}")

            rows.append(row)
            progress.advance(task)
            time.sleep(DELAY_SECONDS)

    return rows


def write_output(path: Path, rows: List[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    suffix = path.suffix.lower()

    if suffix == ".csv":
        with path.open("w", encoding="utf-8", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=OUTPUT_FIELDS, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(rows)
        return

    if suffix in {".xlsx", ".xlsm"}:
        wb = Workbook()
        ws = wb.active
        ws.title = "catalogo_win"
        ws.append(OUTPUT_FIELDS)
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row in rows:
            ws.append([row.get(field, "") for field in OUTPUT_FIELDS])
        for column_cells in ws.columns:
            header = str(column_cells[0].value or "")
            max_len = max(len(str(cell.value or "")) for cell in column_cells[:200])
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, len(header) + 2), 55)
        wb.save(path)
        return

    with path.open("w", encoding="utf-8") as fh:
        json.dump({"resultados": rows}, fh, indent=2, ensure_ascii=False)


def print_summary(rows: List[Dict[str, Any]], output_path: Path) -> None:
    counts: Dict[str, int] = {}
    for row in rows:
        counts[row["status"]] = counts.get(row["status"], 0) + 1

    table = Table(title="Resumen scraper catalogo.win")
    table.add_column("Estado", style="cyan")
    table.add_column("Cantidad", justify="right")
    for status, count in sorted(counts.items()):
        table.add_row(status, str(count))
    console.print(table)
    console.print(f"[bold green]Salida escrita:[/bold green] {output_path}")


def main() -> int:
    args = parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output)

    try:
        raw = load_json(input_path)
        items = unwrap_items(raw)
        rows = scrape_items(items, args.key)
        write_output(output_path, rows)
        print_summary(rows, output_path)
        return 0
    except CriticalScraperError as exc:
        console.print(f"[bold red]Error critico:[/bold red] {exc}")
        return 1
    except OSError as exc:
        console.print(f"[bold red]Error de archivo:[/bold red] {exc}")
        return 1
    except KeyboardInterrupt:
        console.print("[bold yellow]Ejecucion interrumpida por el usuario.[/bold yellow]")
        return 1


if __name__ == "__main__":
    sys.exit(main())
