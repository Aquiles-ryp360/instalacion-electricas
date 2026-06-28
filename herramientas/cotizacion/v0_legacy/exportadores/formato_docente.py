#!/usr/bin/env python3
"""Exporta resultados de catalogo.win a un Excel presentable para revision docente."""

from __future__ import annotations

import argparse
import csv
import json
import sys
from datetime import date
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


RAW_FIELDS = [
    "index",
    "input_key",
    "query",
    "query_usada",
    "query_intentos",
    "codigo",
    "nombre",
    "catalogo_id",
    "codigo_osce",
    "codigo_onu",
    "codigo_searchable",
    "unidad_medida",
    "tipo_bien",
    "precio_ref",
    "nombre_grupo",
    "nombre_clase",
    "nombre_familia",
    "fecha_alta",
    "total_hits",
    "processing_time_ms",
    "status",
    "error",
]

PRESENTATION_FIELDS = [
    "Item",
    "CANT.",
    "UNID.",
    "CODIGO DEL BIEN",
    "DESCRIPCION",
    "MATERIAL DEL BOM",
    "CONSULTA USADA",
    "ESTADO",
    "OBSERVACION",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Genera un Excel de presentacion docente desde resultados del "
            "scraper de catalogo.win."
        )
    )
    parser.add_argument("--input", required=True, help="Resultado del scraper .xlsx, .csv o .json.")
    parser.add_argument("--output", required=True, help="Ruta .xlsx de salida.")
    parser.add_argument("--bom", help="BOM JSON opcional para traer cantidad, unidad y precio.")
    parser.add_argument(
        "--titulo",
        default="REQUERIMIENTO DE INSUMOS PARA INSTALACIONES ELECTRICAS",
        help="Titulo del requerimiento.",
    )
    parser.add_argument(
        "--proyecto",
        default="Instalaciones electricas interiores - vivienda unifamiliar",
        help="Nombre corto del proyecto para la portada.",
    )
    return parser.parse_args()


def load_rows(path: Path) -> List[Dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        headers = [str(cell.value or "").strip() for cell in ws[1]]
        return [
            {headers[i]: value for i, value in enumerate(values) if i < len(headers)}
            for values in ws.iter_rows(min_row=2, values_only=True)
            if any(value is not None and str(value).strip() for value in values)
        ]
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            return list(csv.DictReader(fh))
    with path.open("r", encoding="utf-8") as fh:
        raw = json.load(fh)
    if isinstance(raw, dict):
        rows = raw.get("resultados") or raw.get("rows") or raw.get("data") or []
    else:
        rows = raw
    if not isinstance(rows, list):
        raise ValueError("El JSON de entrada debe contener una lista de resultados.")
    return [row for row in rows if isinstance(row, dict)]


def load_bom(path: Optional[Path]) -> Dict[int, Dict[str, Any]]:
    if not path:
        return {}
    with path.open("r", encoding="utf-8") as fh:
        raw = json.load(fh)
    materiales = raw.get("materiales") if isinstance(raw, dict) else raw
    if not isinstance(materiales, list):
        raise ValueError("El BOM debe ser una lista o un objeto con materiales[].")
    return {
        index: item
        for index, item in enumerate(materiales, start=1)
        if isinstance(item, dict)
    }


def as_int(value: Any, default: int = 0) -> int:
    try:
        if value is None or value == "":
            return default
        return int(float(value))
    except Exception:
        return default


def text(value: Any) -> str:
    return "" if value is None else str(value)


def enrich_rows(rows: Iterable[Dict[str, Any]], bom_by_index: Dict[int, Dict[str, Any]]) -> List[Dict[str, Any]]:
    enriched: List[Dict[str, Any]] = []
    for fallback_index, row in enumerate(rows, start=1):
        index = as_int(row.get("index"), fallback_index)
        bom = bom_by_index.get(index, {})
        cantidad = bom.get("cantidad", "")
        unidad = bom.get("unidad", row.get("unidad_medida", ""))
        material_bom = bom.get("item") or bom.get("nombre") or row.get("query", "")
        observaciones = []
        if row.get("status") != "ok":
            observaciones.append(text(row.get("error") or row.get("status")))
        if row.get("query_usada") and row.get("query_usada") != row.get("query"):
            observaciones.append("Codigo obtenido con consulta alternativa; revisar coincidencia tecnica.")
        if row.get("total_hits") not in ("", None):
            observaciones.append(f"hits={row.get('total_hits')}")

        merged = dict(row)
        merged.update(
            {
                "_item": index,
                "_cantidad": cantidad,
                "_unidad": unidad,
                "_material_bom": material_bom,
                "_observacion": " ".join(obs for obs in observaciones if obs).strip(),
            }
        )
        enriched.append(merged)
    return enriched


def setup_page(ws, title: str) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A22"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4
    ws.title = title[:31]


def apply_common_styles(ws) -> None:
    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
            cell.font = Font(name="Calibri", size=10)


def write_presentation_sheet(wb: Workbook, rows: List[Dict[str, Any]], args: argparse.Namespace) -> None:
    ws = wb.active
    setup_page(ws, "presentacion_docente")

    title_fill = PatternFill("solid", fgColor="D9EAF7")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    section_fill = PatternFill("solid", fgColor="E2F0D9")
    white_bold = Font(name="Calibri", size=10, color="FFFFFF", bold=True)
    bold = Font(name="Calibri", size=10, bold=True)

    ws.merge_cells("A1:I2")
    ws["A1"] = "UNIVERSIDAD NACIONAL DEL ALTIPLANO"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = title_fill

    ws.merge_cells("A3:I3")
    ws["A3"] = "UNIDAD EJECUTORA DE INVERSIONES"
    ws["A3"].font = bold
    ws["A3"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A5:I5")
    ws["A5"] = "CUADRO DE INSUMOS - CODIGOS DE BIEN CATALOGADOS"
    ws["A5"].font = Font(name="Calibri", size=12, bold=True)
    ws["A5"].alignment = Alignment(horizontal="center")
    ws["A5"].fill = title_fill

    ws["A8"] = "PIP / Proyecto"
    ws.merge_cells("B8:I8")
    ws["B8"] = args.proyecto
    ws["A9"] = "Denominacion"
    ws.merge_cells("B9:I9")
    ws["B9"] = args.titulo
    ws["A10"] = "Fecha"
    ws["B10"] = date.today().isoformat()
    ws["D10"] = "Fuente"
    ws.merge_cells("E10:I10")
    ws["E10"] = "catalogo.win + BOM del proyecto"

    ws.merge_cells("A19:I19")
    ws["A19"] = "1.5.1. DESCRIPCION Y CANTIDAD DE LOS BIENES"
    ws["A19"].font = bold
    ws["A19"].fill = section_fill

    start_row = 21
    for col, header in enumerate(PRESENTATION_FIELDS, start=1):
        cell = ws.cell(start_row, col, header)
        cell.fill = header_fill
        cell.font = white_bold
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for offset, row in enumerate(rows, start=1):
        excel_row = start_row + offset
        values = [
            offset,
            row.get("_cantidad", ""),
            row.get("_unidad", ""),
            row.get("codigo", ""),
            row.get("nombre", ""),
            row.get("_material_bom", ""),
            row.get("query_usada", ""),
            row.get("status", ""),
            row.get("_observacion", ""),
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(excel_row, col, value)

    ws.auto_filter.ref = f"A{start_row}:I{start_row + len(rows)}"
    widths = {
        "A": 8,
        "B": 11,
        "C": 12,
        "D": 18,
        "E": 55,
        "F": 42,
        "G": 28,
        "H": 14,
        "I": 38,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    apply_common_styles(ws)


def write_raw_sheet(wb: Workbook, rows: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet("catalogo_win")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Calibri", size=10, color="FFFFFF", bold=True)
    ws.append(RAW_FIELDS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        ws.append([row.get(field, "") for field in RAW_FIELDS])
    ws.auto_filter.ref = f"A1:{get_column_letter(len(RAW_FIELDS))}{len(rows) + 1}"
    for col_idx, field in enumerate(RAW_FIELDS, start=1):
        values = [field] + [text(row.get(field, "")) for row in rows[:100]]
        width = min(max(len(v) for v in values) + 2, 55)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(width, 10)
    apply_common_styles(ws)


def write_workbook(output: Path, rows: List[Dict[str, Any]], args: argparse.Namespace) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    write_presentation_sheet(wb, rows, args)
    write_raw_sheet(wb, rows)
    wb.save(output)


def main() -> int:
    args = parse_args()
    try:
        input_path = Path(args.input)
        output_path = Path(args.output)
        if output_path.suffix.lower() != ".xlsx":
            raise ValueError("La salida del exportador docente debe ser .xlsx.")
        rows = load_rows(input_path)
        bom = load_bom(Path(args.bom)) if args.bom else {}
        enriched = enrich_rows(rows, bom)
        write_workbook(output_path, enriched, args)
        print(f"[ok] Excel de presentacion generado: {output_path}")
        print(f"[ok] Filas exportadas: {len(enriched)}")
        return 0
    except Exception as exc:
        print(f"[error] {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
