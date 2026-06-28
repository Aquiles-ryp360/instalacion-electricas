#!/usr/bin/env python3
"""Scraper de fichas oficiales del buscador publico de PeruCompras."""

from __future__ import annotations

import argparse
from io import BytesIO
import json
import re
import sys
import time
import unicodedata
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    import openpyxl
    import requests
    from bs4 import BeautifulSoup
    from rich.console import Console
    from rich.progress import Progress, SpinnerColumn, TextColumn, TimeElapsedColumn
    from rich.prompt import IntPrompt, Prompt
    from rich.table import Table
except ImportError as exc:  # pragma: no cover - fallo temprano de entorno.
    print(
        "[error] Falta una dependencia. Instala: pip install requests beautifulsoup4 rich openpyxl",
        file=sys.stderr,
    )
    print(f"[error] Modulo faltante: {exc.name}", file=sys.stderr)
    sys.exit(1)


BASE_URL = "https://buscadorcatalogos.perucompras.gob.pe/"
DOWNLOAD_URL = f"{BASE_URL}Search/Download"
REQUEST_TIMEOUT = 30
DEFAULT_DELAY_SECONDS = 0.25
DEFAULT_KEY = "item"
DEFAULT_RETRIES = 3
NO_ENCONTRADO = "NO_ENCONTRADO"

STOPWORDS = {
    "a",
    "al",
    "con",
    "de",
    "del",
    "e",
    "el",
    "en",
    "la",
    "las",
    "los",
    "para",
    "por",
    "tipo",
    "un",
    "una",
    "y",
}

ELECTRICAL_AGREEMENT_MARKERS = (
    "EXT-CE-2024-14",
    "LUMINARIAS",
    "MATERIALES ELECTRICOS",
    "CABLES ELECTRICOS",
)

WRONG_CATEGORY_MARKERS = {
    "computadora",
    "computadoras",
    "desktop",
    "escaner",
    "impresora",
    "laptop",
    "monitor",
    "notebook",
    "portatil",
    "scanner",
}

console = Console()


class CriticalScraperError(RuntimeError):
    """Error que impide continuar con el scraping completo."""


@dataclass(frozen=True)
class MaterialProfile:
    category: str
    section_mm2: str | None = None
    diameter_mm: str | None = None
    watts: str | None = None
    amps: str | None = None
    milliamps: str | None = None
    poles: str | None = None
    inches: str | None = None
    circuits: str | None = None


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Busca fichas-producto de PeruCompras para homologar materiales "
            "electricos de un BOM JSON."
        )
    )
    parser.add_argument("--input", required=True, help="Ruta del JSON de entrada.")
    parser.add_argument("--output", required=True, help="Ruta del JSON de salida.")
    parser.add_argument(
        "--key",
        default=DEFAULT_KEY,
        help="Llave donde esta el nombre del material dentro del BOM. Por defecto: item.",
    )
    parser.add_argument("--limit", type=int, help="Procesa solo los primeros N items.")
    parser.add_argument(
        "--delay",
        type=float,
        default=DEFAULT_DELAY_SECONDS,
        help=f"Segundos de espera entre descargas. Por defecto: {DEFAULT_DELAY_SECONDS}.",
    )
    parser.add_argument(
        "--max-pages",
        type=int,
        default=1,
        help="Compatibilidad legacy; se ignora porque la descarga XLSX no pagina.",
    )
    parser.add_argument(
        "--save-html",
        nargs="?",
        const="",
        help="Compatibilidad legacy; se ignora porque ya no se parsea HTML.",
    )
    parser.add_argument("--verbose", action="store_true", help="Imprime variantes y scores.")
    parser.add_argument(
        "--review-xlsx",
        nargs="?",
        const="AUTO",
        help=(
            "Exporta un Excel de revision con seleccion y alternativas. "
            "Si no se pasa ruta, usa <output>_revision.xlsx."
        ),
    )
    parser.add_argument(
        "--review-menu",
        action="store_true",
        help="Abre un menu interactivo en terminal para revisar resultados al terminar.",
    )
    return parser.parse_args(argv)


def clean_text(value: Any) -> str:
    text = "" if value is None else str(value)
    return re.sub(r"\s+", " ", text).strip()


def remove_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    return "".join(ch for ch in normalized if not unicodedata.combining(ch))


def normalize_material_name(value: str) -> str:
    """Normaliza texto tecnico para busqueda y scoring."""
    text = remove_accents(clean_text(value)).lower()
    replacements = {
        "mm²": "mm2",
        "mm^2": "mm2",
        "m²": "m2",
        "”": '"',
        "“": '"',
        "’": "'",
        "´": "'",
        "`": "'",
        "ø": " diametro ",
        "φ": " diametro ",
        "–": "-",
        "—": "-",
        "×": " x ",
        "*": " x ",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)

    text = re.sub(r'(\d+(?:[./]\d+)?)\s*(?:"|pulgadas?|in\b)', r" \1 pulg ", text)
    text = re.sub(r"\b(\d+)\s*p\b", r" \1 polos ", text)
    text = re.sub(r"\b(\d+)\s*ma\b", r" \1 ma ", text)
    text = re.sub(r"\b(\d+)\s*a\b", r" \1 a ", text)
    text = re.sub(r"\b(\d+)\s*w\b", r" \1 w ", text)
    text = re.sub(r"(\d)\s*mm\s*2\b", r"\1 mm2", text)
    text = re.sub(r"\b(\d+(?:[.,]\d+)?)\s*mm2\b", lambda m: f" {m.group(1).replace(',', '.')} mm2 ", text)
    text = re.sub(r"\b(\d+(?:[.,]\d+)?)\s*mm\b", lambda m: f" {m.group(1).replace(',', '.')} mm ", text)
    text = re.sub(r"[/\\|_+,:;()\[\]{}]", " ", text)
    text = re.sub(r"\s*-\s*", " ", text)
    text = re.sub(r"[^a-z0-9. /-]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def sanitize_search_query(value: str) -> str:
    text = normalize_material_name(value)
    text = re.sub(r"[^0-9a-zA-ZñÑáéíóúÁÉÍÓÚüÜ ._-]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def compact_number(value: str | None) -> str | None:
    if not value:
        return None
    normalized = value.replace(",", ".")
    if "." not in normalized:
        return normalized
    return normalized.rstrip("0").rstrip(".")


def extract_profile(material_name: str) -> MaterialProfile:
    text = normalize_material_name(material_name)
    section = re.search(r"\b(\d+(?:[.]\d+)?)\s*mm2\b", text)
    diameter = re.search(r"\b(\d+(?:[.]\d+)?)\s*mm\b", text)
    watts = re.search(r"\b(\d+(?:[.]\d+)?)\s*w\b", text)
    amps = re.search(r"\b(\d+(?:[.]\d+)?)\s*a\b", text)
    milliamps = re.search(r"\b(\d+(?:[.]\d+)?)\s*ma\b", text)
    poles = re.search(r"\b(\d+)\s*(?:polos|p)\b", text)
    inches = re.search(r"\b(\d+(?:/\d+)?)\s*pulg\b", text)
    circuits = re.search(r"\b(\d+)\s*(?:circuitos|polos)\b", text)

    category = "general"
    if "caja" in text:
        category = "caja"
    elif "diferencial" in text:
        category = "diferencial"
    elif "tablero" in text:
        category = "tablero"
    elif "union" in text and "pvc" in text:
        category = "union_pvc"
    elif "curva" in text and "pvc" in text:
        category = "curva_pvc"
    elif "varilla" in text and "tierra" in text:
        category = "varilla_tierra"
    elif ("conector" in text or "abrazadera" in text) and "tierra" in text:
        category = "conector_tierra"
    elif "puesta" in text and "tierra" in text:
        category = "puesta_tierra"
    elif "cable" in text or "conductor" in text or "thw" in text:
        category = "cable"
    elif "tubo" in text or "tuberia" in text or "pvc" in text:
        category = "tubo_pvc"
    elif "luminaria" in text or "plafon" in text or "led" in text:
        category = "luminaria"
    elif "termomagnetico" in text or "itm" in text:
        category = "interruptor_termomagnetico"
    elif "conmutado" in text or "simple" in text:
        category = "interruptor_control"
    elif "interruptor" in text:
        category = "interruptor_control"
    elif "tomacorriente" in text or "toma corriente" in text or "enchufe" in text:
        category = "tomacorriente"
    elif "cinta" in text and "aisl" in text:
        category = "cinta"
    elif "pararrayo" in text:
        category = "pararrayo"

    return MaterialProfile(
        category=category,
        section_mm2=compact_number(section.group(1)) if section else None,
        diameter_mm=compact_number(diameter.group(1)) if diameter else None,
        watts=compact_number(watts.group(1)) if watts else None,
        amps=compact_number(amps.group(1)) if amps else None,
        milliamps=compact_number(milliamps.group(1)) if milliamps else None,
        poles=poles.group(1) if poles else None,
        inches=inches.group(1) if inches else None,
        circuits=circuits.group(1) if circuits else None,
    )


def token_set(value: str) -> set[str]:
    text = normalize_material_name(value)
    tokens = {token for token in re.findall(r"[a-z0-9.]+", text) if token not in STOPWORDS}
    return {token for token in tokens if len(token) > 1 or token.isdigit()}


def ordered_unique(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        query = sanitize_search_query(value)
        key = normalize_material_name(query)
        if query and key not in seen:
            seen.add(key)
            result.append(query)
    return result


def generate_search_variants(material_name: str) -> list[str]:
    """Genera consultas amplias para no filtrar demasiado desde el buscador."""
    text = normalize_material_name(material_name)
    profile = extract_profile(material_name)
    base = re.sub(r"\b(?:c|circuito)\s*\d+\b", " ", text)
    base = re.sub(r"\b(?:alimentador|general|trifasico|monofasico)\b", " ", base)
    base = re.sub(r"\s+", " ", base).strip()

    spec_tokens = {
        "a",
        "awg",
        "circuitos",
        "ma",
        "mm",
        "mm2",
        "p",
        "polos",
        "pulg",
        "pulgada",
        "pulgadas",
        "w",
    }

    def is_number_token(token: str) -> bool:
        return re.fullmatch(r"\d+(?:\.\d+)?", token) is not None

    base_tokens = [
        token
        for token in re.findall(r"[a-z0-9.]+", base)
        if token not in STOPWORDS and token not in spec_tokens and not is_number_token(token)
    ]
    compact_base = " ".join(base_tokens[:3])

    variants: list[str] = []
    if profile.category == "cable":
        if "thw" in text:
            variants.append("cable thw")
        variants.append("cable electrico")
    elif profile.category == "tubo_pvc":
        variants.extend(["tubo pvc", "tubo pvc sap"])
    elif profile.category == "luminaria":
        variants.extend(["luminaria led", "luminaria"])
    elif profile.category == "interruptor_termomagnetico":
        variants.extend(["interruptor termomagnetico", "interruptor"])
    elif profile.category == "diferencial":
        variants.extend(["interruptor diferencial", "diferencial"])
    elif profile.category == "interruptor_control":
        if "conmutado" in text:
            variants.append("interruptor conmutado")
        elif "simple" in text:
            variants.append("interruptor simple")
        variants.append("interruptor")
    elif profile.category == "tomacorriente":
        variants.extend(["tomacorriente", "tomacorriente doble"])
    elif profile.category == "tablero":
        variants.extend(["tablero electrico", "tablero"])
    elif profile.category == "caja":
        if "octogonal" in text:
            variants.append("caja octogonal")
        elif "estanca" in text or "ip55" in text:
            variants.extend(["caja estanca", "caja pase"])
        elif "registro" in text:
            variants.append("caja registro")
        else:
            variants.append("caja rectangular")
        variants.extend(["caja electrica", "caja pase"])
    elif profile.category == "union_pvc":
        variants.extend(["union pvc", "union pvc sap"])
    elif profile.category == "curva_pvc":
        variants.extend(["curva pvc", "curva pvc sap"])
    elif profile.category == "varilla_tierra":
        variants.extend(["varilla tierra", "puesta tierra"])
    elif profile.category == "conector_tierra":
        variants.extend(["conector tierra", "abrazadera tierra"])
    elif profile.category == "puesta_tierra":
        variants.extend(["puesta tierra", "sistema tierra"])
    elif profile.category == "cinta":
        variants.extend(["cinta aislante", "cinta electrica"])
    elif profile.category == "pararrayo":
        variants.extend(["pararrayo", "sistema pararrayo"])
    elif compact_base:
        variants.append(compact_base)

    if compact_base and len(compact_base.split()) <= 3:
        variants.append(compact_base)

    return ordered_unique(variants)


def load_json(path: Path) -> Any:
    try:
        with path.open("r", encoding="utf-8") as fh:
            return json.load(fh)
    except FileNotFoundError as exc:
        raise CriticalScraperError(f"No existe el archivo de entrada: {path}") from exc
    except json.JSONDecodeError as exc:
        raise CriticalScraperError(f"JSON invalido en {path}: {exc}") from exc
    except OSError as exc:
        raise CriticalScraperError(f"No se pudo leer {path}: {exc}") from exc


def unwrap_items(raw: Any) -> list[dict[str, Any]]:
    if isinstance(raw, list):
        items = raw
    elif isinstance(raw, dict):
        for key in ("materiales", "items", "data", "rows", "resultados"):
            value = raw.get(key)
            if isinstance(value, list):
                items = value
                break
        else:
            items = [raw]
    else:
        raise CriticalScraperError("El JSON debe ser una lista o un objeto con materiales/items/data.")

    normalized: list[dict[str, Any]] = []
    for index, item in enumerate(items, start=1):
        if isinstance(item, dict):
            normalized.append(item)
        else:
            normalized.append({"item": str(item), "_source_index": index})
    return normalized


def get_material_query(item: dict[str, Any], key: str) -> str:
    value = item.get(key)
    if value is None or not str(value).strip():
        raise ValueError(f"El item no tiene valor util en la llave '{key}'.")
    return clean_text(value)


def extract_form_state(html: str) -> dict[str, str]:
    soup = BeautifulSoup(html, "html.parser")
    form = soup.select_one("form#form-search") or soup.select_one("form[method]")
    if not form:
        raise CriticalScraperError("No se encontro el formulario de busqueda en PeruCompras.")

    payload: dict[str, str] = {}
    for field in form.select("input[name]"):
        name = field.get("name")
        if name:
            payload[name] = field.get("value", "")

    token = payload.get("__RequestVerificationToken")
    if not token:
        raise CriticalScraperError("No se encontro __RequestVerificationToken en el formulario.")
    return payload


def request_with_retry(
    session: requests.Session,
    method: str,
    url: str,
    *,
    retries: int = DEFAULT_RETRIES,
    backoff: float = 1.5,
    **kwargs: Any,
) -> requests.Response:
    retry_statuses = {403, 408, 429, 500, 502, 503, 504}
    last_error: Exception | None = None
    for attempt in range(1, retries + 1):
        try:
            response = session.request(method, url, timeout=REQUEST_TIMEOUT, **kwargs)
            if response.status_code in retry_statuses:
                raise requests.HTTPError(f"HTTP {response.status_code}", response=response)
            response.raise_for_status()
            return response
        except (requests.Timeout, requests.ConnectionError, requests.HTTPError) as exc:
            last_error = exc
            if attempt == retries:
                break
            time.sleep(backoff * attempt)
    raise requests.RequestException(f"Fallo {method} {url} tras {retries} intentos: {last_error}")


def bootstrap_session() -> tuple[requests.Session, dict[str, str]]:
    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": (
                "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
                "(KHTML, like Gecko) Chrome/126.0 Safari/537.36"
            ),
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "es-PE,es;q=0.9,en;q=0.8",
        }
    )

    try:
        response = request_with_retry(session, "GET", BASE_URL)
    except requests.RequestException as exc:
        raise CriticalScraperError(f"No se pudo abrir PeruCompras: {exc}") from exc

    if not session.cookies:
        raise CriticalScraperError("El GET inicial no entrego cookies de sesion.")
    return session, extract_form_state(response.text)


def build_search_payload(
    base_payload: dict[str, str],
    query: str,
) -> dict[str, str]:
    payload = dict(base_payload)
    payload.update(
        {
            "From": "Search",
            "IsNewSearch": "True",
            "SearchText": query,
            "SearchTextPrevious": query,
            "Status": "VIGENTE",
            "Pagination.Page": "0",
            "Pagination.Paging": "0",
            "Pagination.LeftMostPage": "0",
        }
    )
    payload.setdefault("ClientFilter.Feature", "[]")
    return payload


EXCEL_COLUMN_ALIASES = {
    "acuerdo_marco": {
        "acuerdo marco",
        "acuerdo",
    },
    "catalogo": {
        "catalogo",
    },
    "categoria": {
        "categoria",
    },
    "nombre_oficial": {
        "descripcion ficha producto",
        "descripcion de ficha producto",
        "ficha producto",
        "descripcion",
        "nombre oficial",
    },
    "marca": {
        "marca",
    },
    "codigo_unico": {
        "nro parte o codigo unico de identificacion",
        "nro parte codigo unico identificacion",
        "numero parte codigo unico identificacion",
        "codigo unico de identificacion",
        "codigo unico",
        "nro parte",
    },
    "url_ficha_pdf": {
        "ficha tecnica",
        "ficha tecnica pdf",
        "url ficha tecnica",
    },
    "url_imagen": {
        "imagen",
        "url imagen",
    },
    "estado_ficha": {
        "estado ficha producto",
        "estado ficha",
        "estado",
    },
}


def normalize_excel_header(value: Any) -> str:
    text = remove_accents(clean_text(value)).lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def canonical_excel_column(header: Any) -> str | None:
    normalized = normalize_excel_header(header)
    if not normalized:
        return None
    for canonical, aliases in EXCEL_COLUMN_ALIASES.items():
        if normalized in aliases:
            return canonical
    return None


def cell_text(row: tuple[Any, ...], header_map: dict[str, int], key: str) -> str:
    index = header_map.get(key)
    if index is None or index >= len(row):
        return ""
    return clean_text(row[index])


def parse_xlsx_cards(content: bytes, *, query: str) -> list[dict[str, Any]]:
    if not content.startswith(b"PK"):
        preview = content[:500].decode("utf-8", errors="replace")
        preview = clean_text(re.sub(r"<[^>]+>", " ", preview))
        raise CriticalScraperError(
            f"El endpoint de descarga no devolvio un XLSX para '{query}'. "
            f"Respuesta: {preview[:180]}"
        )

    try:
        workbook = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise CriticalScraperError(f"No se pudo leer el XLSX descargado para '{query}': {exc}") from exc

    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    header_map: dict[str, int] = {}

    for row in rows:
        detected: dict[str, int] = {}
        for index, header in enumerate(row):
            canonical = canonical_excel_column(header)
            if canonical and canonical not in detected:
                detected[canonical] = index
        if {"nombre_oficial", "categoria", "catalogo"} <= set(detected):
            header_map = detected
            break

    if not header_map:
        workbook.close()
        raise CriticalScraperError(f"No se detectaron cabeceras compatibles en el XLSX para '{query}'.")

    cards: list[dict[str, Any]] = []
    for row in rows:
        if not row or not any(clean_text(value) for value in row):
            continue

        nombre = cell_text(row, header_map, "nombre_oficial")
        if not nombre:
            continue

        cards.append(
            {
                "nombre_oficial": nombre,
                "categoria": cell_text(row, header_map, "categoria"),
                "catalogo": cell_text(row, header_map, "catalogo"),
                "acuerdo_marco": cell_text(row, header_map, "acuerdo_marco"),
                "estado_ficha": cell_text(row, header_map, "estado_ficha"),
                "fecha_publicacion": "",
                "fecha_actualizacion": "",
                "url_imagen": cell_text(row, header_map, "url_imagen"),
                "url_ficha_pdf": cell_text(row, header_map, "url_ficha_pdf"),
                "marca": cell_text(row, header_map, "marca"),
                "codigo_unico": cell_text(row, header_map, "codigo_unico"),
                "caracteristicas": nombre,
            }
        )
    workbook.close()
    return cards


def download_search_cards(
    session: requests.Session,
    base_payload: dict[str, str],
    query: str,
) -> list[dict[str, Any]]:
    headers = {
        "Origin": "https://buscadorcatalogos.perucompras.gob.pe",
        "Referer": BASE_URL,
        "Accept": (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
            "application/vnd.ms-excel,*/*;q=0.8"
        ),
    }
    response = request_with_retry(
        session,
        "POST",
        DOWNLOAD_URL,
        data=build_search_payload(base_payload, query),
        headers=headers,
    )
    return parse_xlsx_cards(response.content, query=query)


def card_text(card: dict[str, Any]) -> str:
    return " ".join(
        clean_text(card.get(key, ""))
        for key in (
            "nombre_oficial",
            "categoria",
            "catalogo",
            "acuerdo_marco",
            "caracteristicas",
        )
    )


def score_card(original_query: str, card: dict[str, Any]) -> dict[str, Any]:
    profile = extract_profile(original_query)
    candidate_text = card_text(card)
    primary_text = " ".join(
        clean_text(card.get(key, ""))
        for key in ("nombre_oficial", "categoria", "catalogo", "acuerdo_marco")
    )
    candidate_norm = normalize_material_name(candidate_text)
    primary_norm = normalize_material_name(primary_text)
    query_tokens = token_set(original_query)
    candidate_tokens = token_set(candidate_text)
    overlap = query_tokens & candidate_tokens
    coverage = len(overlap) / max(len(query_tokens), 1)

    score = coverage * 45
    reasons = [f"cobertura_tokens={coverage:.2f}"]

    agreement_upper = remove_accents(clean_text(card.get("acuerdo_marco", ""))).upper()
    if any(marker in agreement_upper for marker in ELECTRICAL_AGREEMENT_MARKERS):
        score += 8
        reasons.append("acuerdo_electrico")

    wrong_hits = sorted(WRONG_CATEGORY_MARKERS & candidate_tokens)
    if wrong_hits and profile.category not in {"general", "luminaria"}:
        score -= 10
        reasons.append(f"penalizacion_categoria={','.join(wrong_hits)}")

    def add_if(condition: bool, points: float, reason: str) -> None:
        nonlocal score
        if condition:
            score += points
            reasons.append(reason)

    def has_exact(pattern: str) -> bool:
        return re.search(pattern, candidate_norm) is not None

    def has_primary_exact(pattern: str) -> bool:
        return re.search(pattern, primary_norm) is not None

    def has_any(tokens: set[str], *terms: str) -> bool:
        return bool(tokens & set(terms))

    if profile.category == "cable":
        if has_any(candidate_tokens, "tubo", "tubos", "tuberia", "tuberias"):
            score -= 10
            reasons.append("penalizacion_tubo_no_cable")
        add_if(has_any(candidate_tokens, "cable", "cables", "conductor", "conductores"), 12, "tipo_cable")
        add_if(has_any(candidate_tokens, "electrico", "electricos", "electrica", "electricas"), 5, "material_electrico")
        add_if("thw" in candidate_tokens, 5, "thw")
        if profile.section_mm2:
            add_if(
                has_exact(rf"\b{re.escape(profile.section_mm2)}\s*mm2\b"),
                18,
                f"seccion_{profile.section_mm2}_mm2",
            )
    elif profile.category == "tubo_pvc":
        if has_any(candidate_tokens, "cable", "cables", "conductor", "conductores"):
            score -= 18
            reasons.append("penalizacion_cable_no_tubo")
        add_if(has_any(candidate_tokens, "tubo", "tubos", "tuberia", "tuberias"), 12, "tipo_tubo")
        add_if("pvc" in candidate_tokens, 12, "pvc")
        add_if("sap" in candidate_tokens, 6, "sap")
        if profile.diameter_mm:
            add_if(
                has_exact(rf"\b{re.escape(profile.diameter_mm)}\s*mm\b"),
                18,
                f"diametro_{profile.diameter_mm}_mm",
            )
    elif profile.category == "luminaria":
        add_if("luminaria" in candidate_tokens or "luminarias" in candidate_tokens, 14, "tipo_luminaria")
        add_if("led" in candidate_tokens, 10, "led")
        if profile.watts:
            add_if(
                has_exact(rf"\b{re.escape(profile.watts)}\s*w\b"),
                14,
                f"potencia_{profile.watts}_w",
            )
    elif profile.category == "interruptor_termomagnetico":
        if "tomacorriente" in candidate_tokens or "tomacorrientes" in candidate_tokens:
            score -= 10
            reasons.append("penalizacion_tomacorriente_no_interruptor")
        add_if(has_any(candidate_tokens, "interruptor", "interruptores"), 12, "tipo_interruptor")
        add_if("termomagnetico" in candidate_tokens, 18, "termomagnetico")
        if profile.poles:
            add_if(
                has_exact(rf"\b{re.escape(profile.poles)}\s*(?:polos|p)\b"),
                10,
                f"polos_{profile.poles}",
            )
        if profile.amps:
            add_if(
                has_exact(rf"\b{re.escape(profile.amps)}\s*a\b"),
                14,
                f"amperaje_{profile.amps}_a",
            )
    elif profile.category == "diferencial":
        if "termomagnetico" in candidate_tokens:
            score -= 10
            reasons.append("penalizacion_termomagnetico_no_diferencial")
        add_if(has_any(candidate_tokens, "interruptor", "interruptores"), 8, "tipo_interruptor")
        add_if(has_any(candidate_tokens, "diferencial", "diferenciales"), 24, "diferencial")
        if profile.poles:
            add_if(
                has_exact(rf"\b{re.escape(profile.poles)}\s*(?:polos|p)\b"),
                8,
                f"polos_{profile.poles}",
            )
        if profile.amps:
            add_if(
                has_exact(rf"\b{re.escape(profile.amps)}\s*a\b"),
                12,
                f"amperaje_{profile.amps}_a",
            )
        if profile.milliamps:
            add_if(
                has_exact(rf"\b{re.escape(profile.milliamps)}\s*ma\b"),
                10,
                f"sensibilidad_{profile.milliamps}_ma",
            )
    elif profile.category == "interruptor_control":
        if "termomagnetico" in candidate_tokens or "diferencial" in candidate_tokens:
            score -= 18
            reasons.append("penalizacion_proteccion_no_control")
        add_if("interruptor" in candidate_tokens or "interruptores" in candidate_tokens, 18, "tipo_interruptor")
        add_if("simple" in candidate_tokens, 10, "simple")
        add_if("conmutado" in candidate_tokens, 10, "conmutado")
        if profile.amps:
            add_if(
                has_exact(rf"\b{re.escape(profile.amps)}\s*a\b"),
                10,
                f"amperaje_{profile.amps}_a",
            )
    elif profile.category == "tomacorriente":
        add_if(
            "tomacorriente" in candidate_tokens or "tomacorrientes" in candidate_tokens,
            18,
            "tipo_tomacorriente",
        )
        add_if("doble" in candidate_tokens or "duplex" in candidate_tokens, 10, "doble_duplex")
        has_ground = (
            "tierra" in candidate_tokens
            or "puesta" in candidate_tokens
            or re.search(r"\b2p\s+t\b", candidate_norm) is not None
            or re.search(r"\b2\s+polos\s+t\b", candidate_norm) is not None
        )
        add_if(has_ground, 12, "tierra")
        if profile.amps:
            add_if(
                has_primary_exact(rf"\b{re.escape(profile.amps)}\s*a\b"),
                10,
                f"amperaje_{profile.amps}_a",
            )
    elif profile.category == "caja":
        add_if("caja" in candidate_tokens or "cajas" in candidate_tokens, 12, "tipo_caja")
        add_if("rectangular" in candidate_tokens, 8, "rectangular")
        add_if("octogonal" in candidate_tokens, 8, "octogonal")
        add_if("registro" in candidate_tokens, 8, "registro")
        add_if("estanca" in candidate_tokens or "hermetica" in candidate_tokens or "ip55" in candidate_norm, 8, "estanca")
    elif profile.category == "tablero":
        add_if(has_any(candidate_tokens, "tablero", "tableros"), 20, "tipo_tablero")
        add_if(has_any(candidate_tokens, "electrico", "electricos", "electrica", "electricas"), 8, "electrico")
        if profile.circuits:
            add_if(
                has_exact(rf"\b{re.escape(profile.circuits)}\s*(?:circuitos|polos)\b"),
                12,
                f"circuitos_{profile.circuits}",
            )
    elif profile.category in {"union_pvc", "curva_pvc"}:
        expected = "union" if profile.category == "union_pvc" else "curva"
        wrong = "curva" if expected == "union" else "union"
        if wrong in candidate_tokens:
            score -= 15
            reasons.append(f"penalizacion_{wrong}_no_{expected}")
        add_if(expected in candidate_tokens, 18, f"tipo_{expected}")
        add_if("pvc" in candidate_tokens, 12, "pvc")
        add_if("sap" in candidate_tokens, 6, "sap")
        if profile.diameter_mm:
            add_if(
                has_exact(rf"\b{re.escape(profile.diameter_mm)}\s*mm\b"),
                14,
                f"diametro_{profile.diameter_mm}_mm",
            )
    elif profile.category in {"varilla_tierra", "conector_tierra", "puesta_tierra"}:
        add_if("tierra" in candidate_tokens or "puesta" in candidate_tokens, 16, "puesta_tierra")
        add_if("varilla" in candidate_tokens or "electrodo" in candidate_tokens, 12, "varilla_electrodo")
        add_if("conector" in candidate_tokens or "abrazadera" in candidate_tokens, 12, "conector_abrazadera")
        if profile.inches:
            add_if(profile.inches in candidate_norm, 8, f"pulgadas_{profile.inches}")
    elif profile.category == "cinta":
        add_if("cinta" in candidate_tokens, 12, "tipo_cinta")
        add_if("aislante" in candidate_tokens or "aislar" in candidate_tokens, 12, "aislante")
    elif profile.category == "pararrayo":
        add_if("pararrayo" in candidate_tokens or "pararrayos" in candidate_tokens, 18, "tipo_pararrayo")

    score = max(0, min(100, round(score, 2)))
    return {
        "score": score,
        "categoria_detectada": profile.category,
        "tokens_item": sorted(query_tokens),
        "tokens_match": sorted(overlap),
        "razones": reasons,
    }


def confidence_from_scores(best: float, second: float | None) -> tuple[str, bool]:
    margin = best - (second or 0)
    if best >= 75 and margin >= 8:
        return "ALTA", False
    if best >= 75:
        return "MEDIA", True
    if best >= 60 and margin >= 5:
        return "MEDIA", False
    if best >= 60:
        return "MEDIA", True
    return "BAJA", True


def dedupe_scored_cards(cards: list[dict[str, Any]]) -> list[dict[str, Any]]:
    best_by_name: dict[str, dict[str, Any]] = {}
    for card in cards:
        key = normalize_material_name(card.get("nombre_oficial", ""))
        if not key:
            continue
        existing = best_by_name.get(key)
        if existing is None or card.get("score", 0) > existing.get("score", 0):
            best_by_name[key] = card
    return sorted(best_by_name.values(), key=lambda item: item.get("score", 0), reverse=True)


def get_cached_download_cards(
    session: requests.Session,
    base_payload: dict[str, str],
    query: str,
    query_cache: dict[str, list[dict[str, Any]]],
) -> tuple[list[dict[str, Any]], bool]:
    cache_key = normalize_material_name(query)
    if cache_key in query_cache:
        return [dict(card) for card in query_cache[cache_key]], False

    cards = download_search_cards(session, base_payload, query)
    query_cache[cache_key] = [dict(card) for card in cards]
    return cards, True


def search_one(
    session: requests.Session,
    base_payload: dict[str, str],
    index: int,
    item: dict[str, Any],
    args: argparse.Namespace,
    query_cache: dict[str, list[dict[str, Any]]],
) -> dict[str, Any]:
    original_query = get_material_query(item, args.key)
    variants = generate_search_variants(original_query)
    all_cards: list[dict[str, Any]] = []
    variant_errors: list[str] = []

    for variant_index, query in enumerate(variants, start=1):
        try:
            cards, used_network = get_cached_download_cards(
                session,
                base_payload,
                query,
                query_cache=query_cache,
            )
        except (CriticalScraperError, requests.RequestException) as exc:
            variant_errors.append(f"{query}: {exc}")
            if args.verbose:
                console.print(f"[yellow]item {index} variante {variant_index}: {query} fallo; se continua[/]")
            continue
        for card in cards:
            scored = dict(card)
            scoring = score_card(original_query, scored)
            scored.update(scoring)
            scored["query_usada"] = query
            scored["variant_index"] = variant_index
            all_cards.append(scored)
        if args.verbose:
            cache_text = "" if used_network else " cache"
            console.print(f"[dim]item {index} variante {variant_index}: {query} -> {len(cards)} filas XLSX{cache_text}[/]")
        if all_cards:
            current_ranked = dedupe_scored_cards(all_cards)
            current_best = current_ranked[0]
            current_second = current_ranked[1]["score"] if len(current_ranked) > 1 else None
            _, current_review = confidence_from_scores(current_best["score"], current_second)
            if current_best["score"] >= 82 and not current_review:
                break
        if used_network and variant_index < len(variants):
            time.sleep(args.delay)

    ranked = dedupe_scored_cards(all_cards)
    if not ranked:
        return {
            "index": index,
            "input_key": args.key,
            "bom_item_original": item,
            "query_original": original_query,
            "query_variantes": variants,
            "query_usada": "",
            "status": NO_ENCONTRADO,
            "confianza": "BAJA",
            "score": 0,
            "requiere_revision": True,
            "mensaje": (
                "La descarga XLSX no devolvio filas compatibles para las variantes probadas."
                if not variant_errors
                else "Sin candidatos utiles; variantes fallidas: " + " | ".join(variant_errors[:3])
            ),
            "nombre_oficial": "",
            "mejor_resultado": None,
            "candidatos": [],
            "html_debug": [],
            "error": " | ".join(variant_errors),
        }

    best = ranked[0]
    second_score = ranked[1]["score"] if len(ranked) > 1 else None
    confianza, requiere_revision = confidence_from_scores(best["score"], second_score)
    if second_score is not None and best["score"] - second_score < 5:
        requiere_revision = True
    if best["score"] < 35:
        return {
            "index": index,
            "input_key": args.key,
            "bom_item_original": item,
            "query_original": original_query,
            "query_variantes": variants,
            "query_usada": best.get("query_usada", ""),
            "status": NO_ENCONTRADO,
            "confianza": "BAJA",
            "score": best["score"],
            "requiere_revision": True,
            "mensaje": (
                "PeruCompras devolvio filas XLSX, pero ninguna alcanzo el "
                "score minimo de coincidencia tecnica."
            ),
            "nombre_oficial": "",
            "mejor_resultado": best,
            "candidatos": ranked[:10],
            "html_debug": [],
            "error": " | ".join(variant_errors),
        }

    return {
        "index": index,
        "input_key": args.key,
        "bom_item_original": item,
        "query_original": original_query,
        "query_variantes": variants,
        "query_usada": best.get("query_usada", ""),
        "status": "OK",
        "confianza": confianza,
        "score": best["score"],
        "requiere_revision": requiere_revision,
        "mensaje": f"Se evaluaron {len(ranked)} candidatos descargados por XLSX.",
        "nombre_oficial": best["nombre_oficial"],
        "mejor_resultado": best,
        "candidatos": ranked[:10],
        "html_debug": [],
        "error": " | ".join(variant_errors),
    }


def write_output(path: Path, payload: dict[str, Any]) -> None:
    if path.suffix.lower() != ".json":
        raise CriticalScraperError("La salida debe ser un archivo .json.")
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False, indent=2)
        fh.write("\n")


def render_summary(results: list[dict[str, Any]]) -> None:
    total = len(results)
    status_counts = Counter(row.get("status", "ERROR") for row in results)
    confidence_counts = Counter(row.get("confianza", "") for row in results if row.get("status") == "OK")
    review = sum(1 for row in results if row.get("requiere_revision"))

    table = Table(title="Resumen PeruCompras")
    table.add_column("Total", justify="right")
    table.add_column("OK", justify="right", style="green")
    table.add_column("NO_ENCONTRADO", justify="right", style="yellow")
    table.add_column("ERROR", justify="right", style="red")
    table.add_column("Revision", justify="right", style="magenta")
    table.add_row(
        str(total),
        str(status_counts.get("OK", 0)),
        str(status_counts.get(NO_ENCONTRADO, 0)),
        str(status_counts.get("ERROR", 0)),
        str(review),
    )
    console.print(table)
    console.print(
        "[cyan]Confianza:[/] "
        f"ALTA={confidence_counts.get('ALTA', 0)} "
        f"MEDIA={confidence_counts.get('MEDIA', 0)} "
        f"BAJA={confidence_counts.get('BAJA', 0)}"
    )


def resolve_review_xlsx_path(output_path: Path, review_xlsx: str | None) -> Path | None:
    if not review_xlsx:
        return None
    if review_xlsx == "AUTO":
        return output_path.with_name(f"{output_path.stem}_revision.xlsx")
    return Path(review_xlsx)


def excel_value(value: Any) -> str:
    if isinstance(value, (dict, list)):
        text = json.dumps(value, ensure_ascii=False)
    else:
        text = clean_text(value)
    if text.startswith(("=", "+", "-", "@")):
        text = f"'{text}"
    return text[:32000]


def style_review_sheet(worksheet: Any) -> None:
    header_fill = openpyxl.styles.PatternFill("solid", fgColor="1F4E78")
    header_font = openpyxl.styles.Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for column in worksheet.columns:
        letter = openpyxl.utils.get_column_letter(column[0].column)
        width = max(len(excel_value(cell.value)) for cell in column)
        worksheet.column_dimensions[letter].width = min(max(width + 2, 10), 70)


def export_review_xlsx(path: Path, payload: dict[str, Any]) -> None:
    results = payload.get("resultados", [])
    workbook = openpyxl.Workbook()

    resumen = workbook.active
    resumen.title = "Resumen"
    resumen.append(["campo", "valor"])
    for key, value in payload.get("resumen", {}).items():
        resumen.append([key, value])
    resumen.append(["", ""])
    for key, value in payload.get("metadata", {}).items():
        resumen.append([key, excel_value(value)])
    style_review_sheet(resumen)

    seleccion = workbook.create_sheet("Seleccion")
    seleccion_headers = [
        "index",
        "codigo",
        "item",
        "cantidad",
        "unidad",
        "uso",
        "status",
        "confianza",
        "score",
        "requiere_revision",
        "query_usada",
        "nombre_oficial",
        "categoria",
        "catalogo",
        "acuerdo_marco",
        "estado_ficha",
        "url_ficha_pdf",
        "razones",
        "tokens_match",
        "mensaje",
        "error",
    ]
    seleccion.append(seleccion_headers)
    for row in results:
        bom = row.get("bom_item_original") or {}
        best = row.get("mejor_resultado") or {}
        seleccion.append(
            [
                row.get("index"),
                bom.get("codigo", ""),
                row.get("query_original", ""),
                bom.get("cantidad", ""),
                bom.get("unidad", ""),
                bom.get("uso", ""),
                row.get("status", ""),
                row.get("confianza", ""),
                row.get("score", ""),
                row.get("requiere_revision", ""),
                row.get("query_usada", ""),
                best.get("nombre_oficial", row.get("nombre_oficial", "")),
                best.get("categoria", ""),
                best.get("catalogo", ""),
                best.get("acuerdo_marco", ""),
                best.get("estado_ficha", ""),
                best.get("url_ficha_pdf", ""),
                ", ".join(best.get("razones", [])),
                ", ".join(best.get("tokens_match", [])),
                row.get("mensaje", ""),
                row.get("error", ""),
            ]
        )
    style_review_sheet(seleccion)

    alternativas = workbook.create_sheet("Alternativas")
    alternativas_headers = [
        "index",
        "rank",
        "item",
        "status_item",
        "confianza_item",
        "score_item",
        "query_usada",
        "score_candidato",
        "categoria_detectada",
        "tokens_match",
        "razones",
        "nombre_oficial",
        "categoria",
        "catalogo",
        "acuerdo_marco",
        "estado_ficha",
        "url_ficha_pdf",
        "url_imagen",
        "marca",
        "codigo_unico",
    ]
    alternativas.append(alternativas_headers)
    for row in results:
        for rank, candidate in enumerate(row.get("candidatos", []), start=1):
            alternativas.append(
                [
                    row.get("index"),
                    rank,
                    row.get("query_original", ""),
                    row.get("status", ""),
                    row.get("confianza", ""),
                    row.get("score", ""),
                    candidate.get("query_usada", ""),
                    candidate.get("score", ""),
                    candidate.get("categoria_detectada", ""),
                    ", ".join(candidate.get("tokens_match", [])),
                    ", ".join(candidate.get("razones", [])),
                    candidate.get("nombre_oficial", ""),
                    candidate.get("categoria", ""),
                    candidate.get("catalogo", ""),
                    candidate.get("acuerdo_marco", ""),
                    candidate.get("estado_ficha", ""),
                    candidate.get("url_ficha_pdf", ""),
                    candidate.get("url_imagen", ""),
                    candidate.get("marca", ""),
                    candidate.get("codigo_unico", ""),
                ]
            )
    style_review_sheet(alternativas)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()


def render_review_table(rows: list[dict[str, Any]], title: str, limit: int = 25) -> None:
    table = Table(title=title)
    table.add_column("#", justify="right", width=4)
    table.add_column("Status", width=14)
    table.add_column("Conf.", width=6)
    table.add_column("Score", justify="right", width=7)
    table.add_column("Rev.", justify="center", width=5)
    table.add_column("Item", overflow="fold")
    table.add_column("Elegido", overflow="fold")
    for row in rows[:limit]:
        table.add_row(
            str(row.get("index", "")),
            row.get("status", ""),
            row.get("confianza", ""),
            str(row.get("score", "")),
            "SI" if row.get("requiere_revision") else "NO",
            row.get("query_original", ""),
            row.get("nombre_oficial", "")[:140],
        )
    console.print(table)
    if len(rows) > limit:
        console.print(f"[dim]Mostrando {limit} de {len(rows)} filas.[/]")


def render_item_detail(results: list[dict[str, Any]], item_index: int) -> None:
    row = next((item for item in results if item.get("index") == item_index), None)
    if row is None:
        console.print(f"[yellow]No existe item con index {item_index}.[/]")
        return

    console.rule(f"[bold cyan]Detalle item {item_index}")
    detail = Table(show_header=False)
    detail.add_column("Campo", style="cyan", width=18)
    detail.add_column("Valor", overflow="fold")
    for key in ("query_original", "status", "confianza", "score", "requiere_revision", "query_usada", "mensaje", "error"):
        detail.add_row(key, excel_value(row.get(key, "")))
    console.print(detail)

    candidates = row.get("candidatos", [])
    if not candidates:
        console.print("[yellow]Sin alternativas registradas.[/]")
        return

    table = Table(title="Top alternativas")
    table.add_column("#", justify="right", width=3)
    table.add_column("Score", justify="right", width=7)
    table.add_column("Query", width=22)
    table.add_column("Categoria", width=24)
    table.add_column("Nombre oficial", overflow="fold")
    table.add_column("Razones", overflow="fold")
    for rank, candidate in enumerate(candidates[:10], start=1):
        table.add_row(
            str(rank),
            str(candidate.get("score", "")),
            candidate.get("query_usada", ""),
            candidate.get("categoria", ""),
            candidate.get("nombre_oficial", "")[:180],
            ", ".join(candidate.get("razones", []))[:180],
        )
    console.print(table)


def interactive_review_menu(results: list[dict[str, Any]]) -> None:
    while True:
        console.rule("[bold cyan]Menu de revision PeruCompras")
        console.print("[1] Ver todos")
        console.print("[2] Ver requieren revision")
        console.print("[3] Ver NO_ENCONTRADO")
        console.print("[4] Ver confianza MEDIA")
        console.print("[5] Ver confianza BAJA")
        console.print("[6] Ver detalle por index")
        console.print("[0] Salir")
        choice = Prompt.ask("Opcion", choices=["0", "1", "2", "3", "4", "5", "6"], default="2")

        if choice == "0":
            break
        if choice == "1":
            render_review_table(results, "Todos los resultados")
        elif choice == "2":
            render_review_table([row for row in results if row.get("requiere_revision")], "Requieren revision")
        elif choice == "3":
            render_review_table([row for row in results if row.get("status") == NO_ENCONTRADO], NO_ENCONTRADO)
        elif choice == "4":
            render_review_table([row for row in results if row.get("confianza") == "MEDIA"], "Confianza MEDIA")
        elif choice == "5":
            render_review_table([row for row in results if row.get("confianza") == "BAJA"], "Confianza BAJA")
        elif choice == "6":
            item_index = IntPrompt.ask("Index del item")
            render_item_detail(results, item_index)


def run(args: argparse.Namespace) -> int:
    if args.limit is not None and args.limit <= 0:
        raise CriticalScraperError("--limit debe ser mayor a 0.")
    if args.delay < 0:
        raise CriticalScraperError("--delay no puede ser negativo.")

    input_path = Path(args.input)
    output_path = Path(args.output)
    review_xlsx_path = resolve_review_xlsx_path(output_path, args.review_xlsx)
    items = unwrap_items(load_json(input_path))
    if args.limit is not None:
        items = items[: args.limit]

    console.rule("[bold cyan]Scraper PeruCompras - fichas oficiales")
    console.print(f"[cyan]Input:[/] {input_path}")
    console.print(f"[cyan]Output:[/] {output_path}")
    console.print(f"[cyan]Items:[/] {len(items)}")
    console.print(f"[cyan]Endpoint:[/] {DOWNLOAD_URL}")
    console.print(f"[cyan]Modo:[/] descarga XLSX en memoria  [cyan]Delay:[/] {args.delay}s")
    if review_xlsx_path:
        console.print(f"[cyan]Excel revision:[/] {review_xlsx_path}")
    if args.review_menu:
        console.print("[cyan]Menu revision:[/] se abrira al terminar.")
    if args.max_pages != 1 or args.save_html:
        console.print("[yellow]Nota:[/] --max-pages y --save-html se ignoran en modo XLSX.")

    session, base_payload = bootstrap_session()
    console.print("[green]Sesion iniciada; cookies y token anti-forgery capturados.[/]")

    results: list[dict[str, Any]] = []
    query_cache: dict[str, list[dict[str, Any]]] = {}
    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        TimeElapsedColumn(),
        console=console,
    ) as progress:
        task = progress.add_task("Buscando fichas", total=len(items))
        for index, item in enumerate(items, start=1):
            try:
                row = search_one(session, base_payload, index, item, args, query_cache)
                if row["status"] == "OK":
                    console.print(
                        f"[green]OK[/] {index}: {row['query_original']} -> "
                        f"{row['confianza']} score={row['score']} "
                        f"rev={row['requiere_revision']}"
                    )
                else:
                    console.print(f"[yellow]{NO_ENCONTRADO}[/] {index}: {row['query_original']}")
            except CriticalScraperError as exc:
                console.print(f"[yellow]Sesion reiniciada por advertencia:[/] {exc}")
                try:
                    session, base_payload = bootstrap_session()
                    row = search_one(session, base_payload, index, item, args, query_cache)
                except Exception as retry_exc:
                    row = error_row(index, item, args.key, retry_exc)
                    console.print(f"[yellow]Advertencia[/] item {index}: {retry_exc}")
            except Exception as exc:
                row = error_row(index, item, args.key, exc)
                console.print(f"[yellow]Advertencia[/] item {index}: {exc}")
            results.append(row)
            progress.advance(task)

    output_payload = {
        "metadata": {
            "fuente": BASE_URL,
            "endpoint_download": DOWNLOAD_URL,
            "tipo_fuente": "ficha_producto_homologacion",
            "nota": "PeruCompras se usa para homologar fichas-producto; no como fuente de precio final.",
            "generado_en": datetime.now(timezone.utc).isoformat(),
            "input": str(input_path),
            "output": str(output_path),
            "key": args.key,
            "limit": args.limit,
            "delay_seconds": args.delay,
            "modo_busqueda": "xlsx_download_in_memory",
            "max_pages_legacy_ignorado": args.max_pages,
            "save_html_legacy_ignorado": args.save_html,
            "review_xlsx": str(review_xlsx_path) if review_xlsx_path else "",
            "review_menu": bool(args.review_menu),
            "total_items": len(items),
            "consultas_cacheadas": len(query_cache),
            "status_default": "VIGENTE",
        },
        "resumen": {
            "ok": sum(1 for row in results if row.get("status") == "OK"),
            "no_encontrado": sum(1 for row in results if row.get("status") == NO_ENCONTRADO),
            "error": sum(1 for row in results if row.get("status") == "ERROR"),
            "requieren_revision": sum(1 for row in results if row.get("requiere_revision")),
            "confianza_alta": sum(1 for row in results if row.get("confianza") == "ALTA"),
            "confianza_media": sum(1 for row in results if row.get("confianza") == "MEDIA"),
            "confianza_baja": sum(1 for row in results if row.get("confianza") == "BAJA"),
        },
        "resultados": results,
    }
    write_output(output_path, output_payload)
    if review_xlsx_path:
        export_review_xlsx(review_xlsx_path, output_payload)
    render_summary(results)
    console.print(f"[green]JSON guardado:[/] {output_path}")
    if review_xlsx_path:
        console.print(f"[green]Excel de revision guardado:[/] {review_xlsx_path}")
    if args.review_menu:
        interactive_review_menu(results)
    return 0


def error_row(index: int, item: dict[str, Any], key: str, exc: Exception) -> dict[str, Any]:
    return {
        "index": index,
        "input_key": key,
        "bom_item_original": item,
        "query_original": clean_text(item.get(key, "")) if isinstance(item, dict) else "",
        "query_variantes": [],
        "query_usada": "",
        "status": "ERROR",
        "confianza": "BAJA",
        "score": 0,
        "requiere_revision": True,
        "mensaje": "",
        "nombre_oficial": "",
        "mejor_resultado": None,
        "candidatos": [],
        "html_debug": [],
        "error": str(exc),
    }


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    try:
        return run(args)
    except CriticalScraperError as exc:
        console.print(f"[red]Error critico:[/] {exc}", stderr=True)
        return 1
    except requests.RequestException as exc:
        console.print(f"[red]Error de red:[/] {exc}", stderr=True)
        return 1
    except KeyboardInterrupt:
        console.print("[yellow]Ejecucion interrumpida por el usuario.[/]", stderr=True)
        return 1


if __name__ == "__main__":
    sys.exit(main())
