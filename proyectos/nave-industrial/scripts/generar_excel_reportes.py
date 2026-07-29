#!/usr/bin/env python3
"""Generador de libros Excel (.xlsx) formateados y profesionales para Nave Industrial.

Salidas:
  1. entregables/cuadro_de_cargas_nave_industrial.xlsx
  2. entregables/metrados_y_presupuesto_nave_industrial.xlsx
"""

import json
import os
from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[3]
INPUT_JSON = REPO_ROOT / "proyectos" / "nave-industrial" / "diseno-electrico" / "datos" / "cargas-industriales.json"
ENTREGABLES_DIR = REPO_ROOT / "proyectos" / "nave-industrial" / "entregables"

# Estilos de openpyxl
FONT_TITLE = Font(name="Calibri", size=16, bold=True, color="1F4E79")
FONT_SUBTITLE = Font(name="Calibri", size=11, italic=True, color="595959")
FONT_SECTION = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BOLD = Font(name="Calibri", size=11, bold=True, color="000000")
FONT_NORMAL = Font(name="Calibri", size=11, color="000000")

FILL_SECTION = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
FILL_HEADER = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
FILL_ZEBRA = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
FILL_TOTAL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

BORDER_THIN = Side(border_style="thin", color="D9D9D9")
BORDER_THICK_BOTTOM = Side(border_style="medium", color="1F4E79")
BORDER_DOUBLE_BOTTOM = Side(border_style="double", color="000000")

CELL_BORDER = Border(left=BORDER_THIN, right=BORDER_THIN, top=BORDER_THIN, bottom=BORDER_THIN)
HEADER_BORDER = Border(left=BORDER_THIN, right=BORDER_THIN, top=BORDER_THIN, bottom=BORDER_THICK_BOTTOM)
TOTAL_BORDER = Border(top=BORDER_THIN, bottom=BORDER_DOUBLE_BOTTOM)

ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")


def aplicar_auto_ancho(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if cell.number_format and "S/." in cell.number_format:
                val_str += "    "
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)


def generar_excel_cuadro_cargas(data, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cuadro de Cargas y M.D."

    # Encabezado principal
    ws.merge_cells("A1:K1")
    ws["A1"] = "CUADRO DE CARGAS Y MAXIMA DEMANDA - NAVE INDUSTRIAL 20x40m"
    ws["A1"].font = FONT_TITLE
    ws["A1"].alignment = ALIGN_LEFT

    ws.merge_cells("A2:K2")
    ws["A2"] = f"Propietario: {data.get('propietario', 'Renzo Gabriel Mamani Galindo')} | Ubicacion: Juliaca, Puno | Tension: 380V/220V 3F | Sistema: TN-S"
    ws["A2"].font = FONT_SUBTITLE
    ws["A2"].alignment = ALIGN_LEFT

    # Tabla 1: Resumen de Circuitos
    row = 4
    headers = [
        "Circuito", "Descripcion", "Tipo", "P. Inst. (kW)",
        "Factor Demanda", "M. Demanda (kW)", "F.P.", "Corriente Ib (A)",
        "Seccion (mm2)", "ITM Sugerido", "Caida dV (%)"
    ]

    ws.cell(row=row, column=1, value="1. CUADRO DE CARGAS POR CIRCUITO").font = FONT_SECTION
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_SECTION
        cell.alignment = ALIGN_LEFT
    row += 1

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = HEADER_BORDER
    row += 1

    total_pi = 0.0
    total_md = 0.0
    v = data.get("tension_v", 380)

    for i, c in enumerate(data.get("circuitos", [])):
        pi = c.get("potencia_kw", 0.0)
        fd = c.get("factor_demanda", 1.0)
        md = pi * fd
        fp = c.get("fp", 0.90)
        ib = (md * 1000) / (1.732 * v * fp) if (v and fp and md) else 0.0
        sec = c.get("seccion_mm2", 4)
        long_m = c.get("longitud_m", 20)
        # Caida de tension trifasica %
        dv = (1.732 * long_m * ib * 0.0175) / (sec * v) * 100 if sec else 0.0

        total_pi += pi
        total_md += md

        fill = FILL_ZEBRA if i % 2 == 1 else PatternFill(fill_type=None)

        r_data = [
            c["id"], c["descripcion"], c.get("tipo", "").upper(),
            pi, fd, md, fp, ib, sec, f"3P-{c.get('proteccion_itm_a', 16)}A", dv
        ]

        for col_idx, val in enumerate(r_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = FONT_NORMAL
            cell.fill = fill
            cell.border = CELL_BORDER

            if col_idx in [1, 3, 9, 10]:
                cell.alignment = ALIGN_CENTER
            elif col_idx == 2:
                cell.alignment = ALIGN_LEFT
            else:
                cell.alignment = ALIGN_RIGHT

            if col_idx in [4, 6]:
                cell.number_format = '#,##0.00 "kW"'
            elif col_idx in [5, 7]:
                cell.number_format = '0.00'
            elif col_idx == 8:
                cell.number_format = '#,##0.00 "A"'
            elif col_idx == 11:
                cell.number_format = '0.00 "%"'

        row += 1

    # Fila de Totales
    subtotal_simult = total_md * 0.85
    reserva = subtotal_simult * 0.20
    md_final = subtotal_simult + reserva
    ib_total = (md_final * 1000) / (1.732 * v * 0.90)

    ws.cell(row=row, column=1, value="TOTAL INSTALADO").font = FONT_BOLD
    ws.cell(row=row, column=4, value=total_pi).font = FONT_BOLD
    ws.cell(row=row, column=4).number_format = '#,##0.00 "kW"'
    ws.cell(row=row, column=6, value=total_md).font = FONT_BOLD
    ws.cell(row=row, column=6).number_format = '#,##0.00 "kW"'
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_TOTAL
        cell.border = TOTAL_BORDER
    row += 2

    # Tabla 2: Resumen Ejecutivo de Alimentador
    ws.cell(row=row, column=1, value="2. RESUMEN DE ALIMENTADOR PRINCIPAL Y PROTECCION").font = FONT_SECTION
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    for c in range(1, 7):
        ws.cell(row=row, column=c).fill = FILL_SECTION
    row += 1

    resumen_items = [
        ("Potencia Instalada Total", f"{total_pi:.2f} kW"),
        ("Suma de Maximas Demandas", f"{total_md:.2f} kW"),
        ("Factor de Simultaneidad (CNE-U)", "0.85"),
        ("Subtotal con Simultaneidad", f"{subtotal_simult:.2f} kW"),
        ("Reserva de Crecimiento (20%)", f"{reserva:.2f} kW"),
        ("MAXIMA DEMANDA ADOPTADA", f"{md_final:.2f} kW"),
        ("Corriente de Trabajo (Ib)", f"{ib_total:.2f} A"),
        ("Conductor Alimentador Principal", "50 mm2 N2XH (Ampacidad 102.6A con derrateo)"),
        ("Interruptor Termomagnetico General", "3P - 100A (IP65)"),
        ("Compensacion de Factor de Potencia", "Banco Automatico 15 kVAr (3 x 5 kVAr)"),
        ("Puesta a Tierra (SPAT)", "Malla perimetral 6 varillas Cu 5/8x2.4m (R < 5 ohm)")
    ]

    for item, val in resumen_items:
        c1 = ws.cell(row=row, column=1, value=item)
        c2 = ws.cell(row=row, column=3, value=val)
        c1.font = FONT_BOLD if "MAXIMA" in item else FONT_NORMAL
        c2.font = FONT_BOLD if "MAXIMA" in item else FONT_NORMAL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        row += 1

    aplicar_auto_ancho(ws)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Excel Cuadro de Cargas generado: {output_path}")


def generar_excel_metrados_presupuesto(data, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Presupuesto y Metrados"

    # Encabezado principal
    ws.merge_cells("A1:G1")
    ws["A1"] = "PRESUPUESTO Y METRADOS (NORMA OE.5) - NAVE INDUSTRIAL 20x40m"
    ws["A1"].font = FONT_TITLE

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Propietario: {data.get('propietario', 'Renzo Gabriel Mamani Galindo')} | Proyecto: Instalaciones Eléctricas Industriales"
    ws["A2"].font = FONT_SUBTITLE

    row = 4
    headers = ["Item Partida", "Descripcion de Partida OE.5", "Unidad", "Metrado / Cant.", "Precio Unit. (S/.)", "Parcial (S/.)", "Especificacion / Marca"]

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = HEADER_BORDER
    row += 1

    partidas = [
        ("OE.5.1", "CONEXION A RED EXTERNA Y ACOMETIDA", "", "", "", "", ""),
        ("OE.5.1.1", "Acometida electrica trifasica 380V en cable N2XH 50mm2 (incl. tuberia)", "m", 60, 95.00, "Indeco / PVC SAP"),
        ("OE.5.1.2", "Caja porta-medidor trifasica bimetalica IP65", "und", 1, 450.00, "Schneider Electric"),

        ("OE.5.2", "TABLEROS Y PROTECCIONES ELECTRICAS", "", "", "", "", ""),
        ("OE.5.2.1", "Tablero General TGD 24 polos IP65 con ITM General 100A", "und", 1, 2800.00, "ABB / Schneider"),
        ("OE.5.2.2", "Sub-Tablero de Fuerza TF1 18 polos IP65", "und", 1, 1950.00, "ABB / Schneider"),
        ("OE.5.2.3", "Sub-Tablero de Iluminacion TI1 12 polos IP65", "und", 1, 1400.00, "ABB / Schneider"),
        ("OE.5.2.4", "Interruptores diferenciales superinmunizados (SI) 4P/2P", "glba", 1, 1250.00, "Schneider Electric"),
        ("OE.5.2.5", "Guardamotores ajustables para M1 (10HP) y M2 (15HP)", "jgo", 1, 850.00, "Siemens / ABB"),

        ("OE.5.3", "CANALIZACIONES Y BANDEJAS PORTACABLES", "", "", "", "", ""),
        ("OE.5.3.1", "Bandeja portacables metalica perforada 200x50mm con soportes", "m", 60, 110.00, "Celsa / Bticino"),
        ("OE.5.3.2", "Tuberia PVC SAP 32mm pesada para alimentadores", "m", 30, 22.00, "Pavco Wavin"),
        ("OE.5.3.3", "Tuberia PVC SAP 25mm / 20mm para circuitos derivados", "m", 110, 14.00, "Pavco Wavin"),

        ("OE.5.4", "CONDUCTORES Y CABLES DE ENERGIA", "", "", "", "", ""),
        ("OE.5.4.1", "Cable libre de halogenos N2XH 16mm2 (Maquinaria y BPF)", "m", 40, 38.00, "Indeco"),
        ("OE.5.4.2", "Cable libre de halogenos N2XH 10mm2 (Compresor M2)", "m", 30, 26.00, "Indeco"),
        ("OE.5.4.3", "Cable libre de halogenos N2XH 6mm2 (Grua M1 y Tomas Ind.)", "m", 80, 18.00, "Indeco"),
        ("OE.5.4.4", "Cable libre de halogenos N2XH 4mm2 (Iluminacion y Oficinas)", "m", 100, 12.00, "Indeco"),

        ("OE.5.5", "ALUMBRADO Y TOMACORRIENTES INDUSTRIALES", "", "", "", "", ""),
        ("OE.5.5.1", "Campana LED High-Bay 150W 5000K IP65 (Produccion/Almacen)", "und", 20, 320.00, "Philips / Osram"),
        ("OE.5.5.2", "LED Panel 60x60 40W 4000K para oficinas", "und", 10, 85.00, "Philips"),
        ("OE.5.5.3", "Tomacorriente trifasico Stecker 32A 380V IP67", "und", 4, 180.00, "Mennekes / Scame"),
        ("OE.5.5.4", "Tomacorriente doble Shucko 16A 220V pesados", "und", 8, 45.00, "Bticino Modus"),

        ("OE.5.6", "PUESTA A TIERRA Y COMPENSACION DE FP", "", "", "", "", ""),
        ("OE.5.6.1", "Malla SPAT perimetral con 6 varillas Cu 5/8x2.4m + cable 35mm2", "glba", 1, 3400.00, "Copperweld / Thorjel"),
        ("OE.5.6.2", "Banco de condensadores automatico 15 kVAr (3 x 5 kVAr)", "und", 1, 4800.00, "ABB / Epcos")
    ]

    total_presupuesto = 0.0

    for i, p in enumerate(partidas):
        if len(p) == 7:
            codigo, desc, und, cant, pu, _parcial_dummy, esp = p
        else:
            codigo, desc, und, cant, pu, esp = p
        is_heading = und == ""

        if is_heading:
            row_fill = FILL_SECTION
            row_font = FONT_SECTION
        else:
            row_fill = FILL_ZEBRA if i % 2 == 1 else PatternFill(fill_type=None)
            row_font = FONT_NORMAL

        cell_cod = ws.cell(row=row, column=1, value=codigo)
        cell_desc = ws.cell(row=row, column=2, value=desc)
        cell_und = ws.cell(row=row, column=3, value=und)
        cell_cant = ws.cell(row=row, column=4, value=cant if not is_heading else "")
        cell_pu = ws.cell(row=row, column=5, value=pu if not is_heading else "")

        if not is_heading and isinstance(cant, (int, float)) and isinstance(pu, (int, float)):
            parcial = cant * pu
            total_presupuesto += parcial
            cell_parcial = ws.cell(row=row, column=6, value=parcial)
            cell_parcial.number_format = '"S/." #,##0.00'
        else:
            cell_parcial = ws.cell(row=row, column=6, value="")

        cell_esp = ws.cell(row=row, column=7, value=esp)

        for c in range(1, 8):
            cell = ws.cell(row=row, column=c)
            cell.font = row_font
            cell.fill = row_fill
            cell.border = CELL_BORDER

        cell_cod.alignment = ALIGN_CENTER
        cell_desc.alignment = ALIGN_LEFT
        cell_und.alignment = ALIGN_CENTER
        cell_cant.alignment = ALIGN_RIGHT
        cell_pu.alignment = ALIGN_RIGHT
        cell_parcial.alignment = ALIGN_RIGHT
        cell_esp.alignment = ALIGN_LEFT

        if not is_heading:
            cell_pu.number_format = '"S/." #,##0.00'

        row += 1

    # Fila de Costo Directo y Totales
    ws.cell(row=row, column=2, value="TOTAL COSTO DIRECTO (S/.)").font = FONT_BOLD
    c_tot = ws.cell(row=row, column=6, value=total_presupuesto)
    c_tot.font = FONT_BOLD
    c_tot.number_format = '"S/." #,##0.00'

    for c in range(1, 8):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_TOTAL
        cell.border = TOTAL_BORDER
    row += 1

    igv = total_presupuesto * 0.18
    total_con_igv = total_presupuesto + igv

    ws.cell(row=row, column=2, value="IGV (18%)").font = FONT_BOLD
    c_igv = ws.cell(row=row, column=6, value=igv)
    c_igv.font = FONT_BOLD
    c_igv.number_format = '"S/." #,##0.00'
    row += 1

    ws.cell(row=row, column=2, value="PRESUPUESTO TOTAL GENERAL (S/.)").font = FONT_BOLD
    c_gen = ws.cell(row=row, column=6, value=total_con_igv)
    c_gen.font = FONT_BOLD
    c_gen.number_format = '"S/." #,##0.00'

    aplicar_auto_ancho(ws)
    wb.save(output_path)
    print(f"Excel Metrados y Presupuesto generado: {output_path}")


def main():
    if not INPUT_JSON.exists():
        print(f"ERROR: No existe {INPUT_JSON}")
        return
    with open(INPUT_JSON, encoding="utf-8") as f:
        data = json.load(f)

    ENTREGABLES_DIR.mkdir(parents=True, exist_ok=True)
    excel_cargas = ENTREGABLES_DIR / "cuadro_de_cargas_nave_industrial.xlsx"
    excel_presupuesto = ENTREGABLES_DIR / "metrados_y_presupuesto_nave_industrial.xlsx"

    generar_excel_cuadro_cargas(data, excel_cargas)
    generar_excel_metrados_presupuesto(data, excel_presupuesto)


if __name__ == "__main__":
    main()
